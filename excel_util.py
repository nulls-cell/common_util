from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font
from openpyxl.utils import get_column_letter


class ExcelUtil(Workbook):

    # 将每行的数据转化，如果是数据格式保留，其他转化成字符串
    @staticmethod
    def set_row(row):
        assert(isinstance(row, list) or isinstance(row, tuple))
        _row = []
        for i in range(len(row)):
            if str(row[i]).replace('.', '').isdigit() and str(row[i]).split('.')[0].__len__() <= 13:
                _row.append(eval(str(row[i])))
            else:
                _row.append(str(row[i]))
        return _row

    # 指定sheet_name写入数据，如果sheet_name不存在，则先创建，再写入
    def __write_sheet(self, _sheet_name, head_data, result_data):
        if _sheet_name not in self.sheetnames:
            self.create_sheet(_sheet_name)
        ws = self[_sheet_name]
        ws.append(head_data)
        for row in result_data:
            ws.append(ExcelUtil.set_row(row))

    # 写入excel，如果sheet_name不存在，创建并写入，支持传入单页最大的写入行数
    def write_excel(self, _sheet_name, head_data, result_data, page_max=1048576):
        row_count = len(result_data)  # 数据行数
        page_count = max(row_count // page_max, 1)   # 一共需要的页数
        # 如果1页的话，直接写入到当前sheet
        if page_count == 1:
            self.__write_sheet(_sheet_name, head_data, result_data)
        # 如果1页写不下，则分多页写入
        else:
            for i in range(page_count):
                tem_sheet_name = _sheet_name + str(i+1)
                self.__write_sheet(tem_sheet_name, head_data, result_data)

    # 设置格式（设置字体为微软雅黑，设置字体首行加粗）
    def set_style(self, _sheet_name):
        font = Font(name=u'微软雅黑', bold=False)
        font_bold = Font(name=u'微软雅黑', bold=True)
        thin = Side(border_style=u"thin", color=u"000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        ws = self[_sheet_name]
        row, column = ws.max_row, ws.max_column
        for i in range(1, row + 1):
            for j in range(1, column + 1):
                ws.cell(column=j, row=i).border = border
                if i == 1:
                    ws.cell(column=j, row=i).font = font_bold
                else:
                    ws.cell(column=j, row=i).font = font

    # 设置列宽（设置列宽符合当列最大长度）
    def set_width(self, _sheet_name):
        ws = self[_sheet_name]
        row = ws.max_row
        column = ws.max_column
        width_list = []
        for j in range(1, column + 1):
            max_len = 0
            for i in range(1, row + 1):
                tem_len = float(len((str(ws.cell(column=j, row=i).value).encode('utf-8'))) +
                                len(str(ws.cell(column=j, row=i).value)) * 3) / 4
                max_len = max(max_len, tem_len)
            width_list.append(max_len)
        for i in range(len(width_list)):
            col = str(get_column_letter(i + 1))
            ws.column_dimensions[col].width = width_list[i] + 5

    # 调整excel格式（包括设置字体和列宽）
    def set_fit(self):
        for sh_name in self.sheetnames:
            ws = self[sh_name]
            row, col = ws.max_row, ws.max_column
            if row + col == 2 and ws['A1'].value is None:
                self.remove(ws)
            else:
                self.set_style(sh_name)
                self.set_width(sh_name)

    # 保存excel，继承了父类的方法
    def save(self, filename):
        super(ExcelUtil, self).save(filename=filename)


if __name__ == '__main__':
    # 创建excel对象
    wb = ExcelUtil()
    # 创建表头数据列表
    _head_data = list(range(10))
    # 创建内容数据列表
    _result_data = [list(range(11, 20)), list(range(21, 30))]
    # 写入excel
    wb.write_excel('数字列表', _head_data, _result_data)
    # 调整excel格式，包括字体和列宽
    wb.set_fit()
    # 保存excel
    wb.save(filename='E:/测试数据.xlsx')
